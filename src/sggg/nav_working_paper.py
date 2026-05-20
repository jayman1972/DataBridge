"""Read EHP NAV estimates from the daily NAV Review working paper Excel file."""

from __future__ import annotations

import os
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Spreadsheet label (column B) -> Diamond fund parent GUID
WORKING_PAPER_FUND_IDS: Dict[str, str] = {
    "SUM EHP ALPHA": "415a3530-3034-4536-4432-303030364337",
    "SUM EHP SELECT ALT": "41323030-3031-4144-3637-303030364338",
    "SUM EHP STRAT INC ALT": "41010000-7F2A-D7E8-776F-45484608D91C",
    "SUM EHP TACT GROWTH ALT": "41010000-7F7A-0A65-D559-45484608DB40",
    "SUM EXPON BAL GROW FUND": "01010000-801A-4995-8370-45484608DE57",
}

_NAV_REVIEW_RE = re.compile(
    r"NAV\s+Review\s+(\d{1,2})\.(\d{1,2})\.(\d{4})",
    re.IGNORECASE,
)


def _norm_label(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().upper())


def prior_business_day(d: date) -> date:
    d = d - timedelta(days=1)
    while d.weekday() >= 5:
        d = d - timedelta(days=1)
    return d


def _parse_nav_review_date(path: Path) -> Optional[date]:
    m = _NAV_REVIEW_RE.search(path.stem)
    if not m:
        return None
    month, day, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _parse_money_cell(raw: Any) -> Optional[float]:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if not s or s in ("-", "—"):
        return None
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    s = s.replace(",", "").replace("$", "").strip()
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if neg else v


def find_nav_review_workbook(
    valuation_date: date,
    root: Optional[str] = None,
) -> Tuple[Optional[Path], Optional[str]]:
    """
    Locate the NAV Review workbook for the valuation date.

    Prefer files whose filename date matches valuation_date; otherwise the newest
    NAV Review*.xlsx under root by modification time (on or before valuation_date).
    """
    base = Path(root or os.environ.get("NAV_WORKING_PAPER_ROOT", r"P:\03. Reporting\01. NAV Working Paper Support"))
    if not base.exists():
        return None, f"NAV working paper root not found: {base}"

    candidates: List[Tuple[date, float, Path]] = []
    for path in base.rglob("NAV Review*.xlsx"):
        if path.name.startswith("~$"):
            continue
        file_date = _parse_nav_review_date(path)
        try:
            mtime = path.stat().st_mtime
        except OSError:
            continue
        if file_date and file_date == valuation_date:
            return path, None
        if file_date and file_date <= valuation_date:
            candidates.append((file_date, mtime, path))

    if candidates:
        candidates.sort(key=lambda t: (t[0], t[1]), reverse=True)
        return candidates[0][2], None

    # Fallback: newest file by mtime (any date)
    by_mtime: List[Tuple[float, Path]] = []
    for path in base.rglob("NAV Review*.xlsx"):
        if path.name.startswith("~$"):
            continue
        try:
            by_mtime.append((path.stat().st_mtime, path))
        except OSError:
            continue
    if not by_mtime:
        return None, f"No NAV Review*.xlsx found under {base}"
    by_mtime.sort(reverse=True)
    return by_mtime[0][1], "Used latest NAV Review file by modified time (no exact date match)"


def _rows_from_openpyxl(workbook_path: Path, sheet_name: str, cell_range: str) -> List[List[Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            match = next((n for n in wb.sheetnames if n.upper() == sheet_name.upper()), None)
            if not match:
                raise RuntimeError(f"Sheet {sheet_name!r} not found in {workbook_path.name}")
            sheet_name = match
        ws = wb[sheet_name]
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(cell_range)
        return [
            list(row)
            for row in ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
        ]
    finally:
        wb.close()


def _rows_from_stdlib(workbook_path: Path, sheet_name: str, cell_range: str) -> List[List[Any]]:
    from sggg.xlsx_stdlib import read_sheet_range

    return read_sheet_range(workbook_path, sheet_name, cell_range)


def read_pnl_estimates(
    workbook_path: Path,
    sheet_name: str = "PNL",
    cell_range: str = "A5:C17",
) -> Dict[str, Dict[str, Any]]:
    """
    Read fund labels (column B) and NAV change estimates in dollars (column C).
    Returns dict keyed by normalized spreadsheet label.

    Uses openpyxl when installed; otherwise reads .xlsx via stdlib only (no pip).
    """
    try:
        import openpyxl  # noqa: F401

        grid_rows = _rows_from_openpyxl(workbook_path, sheet_name, cell_range)
    except ImportError:
        grid_rows = _rows_from_stdlib(workbook_path, sheet_name, cell_range)

    out: Dict[str, Dict[str, Any]] = {}
    for row in grid_rows:
        cells = list(row) + [None] * (3 - len(row))
        prior_diff = _parse_money_cell(cells[0])
        label_raw = cells[1]
        estimate = _parse_money_cell(cells[2])
        label = _norm_label(str(label_raw) if label_raw is not None else "")
        if not label or not label.startswith("SUM "):
            continue
        out[label] = {
            "label": str(label_raw).strip() if label_raw else label,
            "estimate_nav_change_dollars": estimate,
            "prior_day_diff": prior_diff,
            "fund_id": WORKING_PAPER_FUND_IDS.get(label),
        }
    return out


def estimates_by_fund_id(
    valuation_date: date,
    root: Optional[str] = None,
) -> Dict[str, Any]:
    """Resolve workbook and return estimates keyed by fund GUID."""
    path, note = find_nav_review_workbook(valuation_date, root=root)
    if not path:
        return {
            "available": False,
            "error": note or "Workbook not found",
            "estimates_by_fund_id": {},
            "estimates_by_label": {},
        }
    try:
        by_label = read_pnl_estimates(path)
    except Exception as e:
        return {
            "available": False,
            "error": str(e),
            "workbook_path": str(path),
            "estimates_by_fund_id": {},
            "estimates_by_label": {},
        }
    by_fund: Dict[str, Dict[str, Any]] = {}
    for label, row in by_label.items():
        fid = row.get("fund_id")
        if fid:
            by_fund[fid] = {**row, "spreadsheet_label": label}
    return {
        "available": True,
        "workbook_path": str(path),
        "note": note,
        "estimates_by_fund_id": by_fund,
        "estimates_by_label": by_label,
    }
