"""Read EHP NAV estimates from the daily Compliance Check workbook (Steps tab, column AB)."""

from __future__ import annotations

import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from sggg.nav_sheet_parse import FUND_NATIVE_CURRENCY, _return_value_to_bps

DEFAULT_ROOT = (
    r"P:\03. Reporting\100. Compliance Check - portfolios (Alt Funds)"
    r"\Daily NI81102 Compliance Check"
)

# Full fund names on Steps tab (column W) -> Diamond fund parent GUID
FUND_NAME_TO_ID: Dict[str, str] = {
    "EHP ALPHA STRATEGIES ALTERNATIVE FUND": "415a3530-3034-4536-4432-303030364337",
    "EHP SELECT ALTERNATIVE FUND": "41323030-3031-4144-3637-303030364338",
    "EHP TACTICAL GROWTH ALTERNATIVE FUND": "41010000-7F7A-0A65-D559-45484608DB40",
    "EHP STRATEGIC INCOME ALTERNATIVE FUND": "41010000-7F2A-D7E8-776F-45484608D91C",
    "EXPONENTIAL BALANCED GROWTH FUND": "01010000-801A-4995-8370-45484608DE57",
}

_COMPLIANCE_DATE_RE = re.compile(
    r"compliance\s+check\s*-\s*(\d{4})\.(\d{2})\.(\d{2})",
    re.IGNORECASE,
)


def _norm_name(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().upper())


def _parse_money(raw: Any) -> Optional[float]:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if not s or s in ("-", "—"):
        return None
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    s = s.replace(",", "").replace("$", "").strip()
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if neg else v


def _parse_workbook_date(path: Path) -> Optional[date]:
    m = _COMPLIANCE_DATE_RE.search(path.stem)
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


def _file_variant(path: Path) -> str:
    name = path.name.lower()
    if "portfolio review-4pm" in name or "portfolio review - 4pm" in name:
        return "4pm"
    return "eod"


def _read_steps_grid_xls(workbook_path: Path, cell_range: str = "W2:AB10") -> List[List[Any]]:
    try:
        import xlrd
    except ImportError as e:
        raise RuntimeError(
            f"Cannot read legacy Excel .xls file {workbook_path.name}: "
            "install xlrd (pip install xlrd). openpyxl only supports .xlsx."
        ) from e

    from sggg.xlsx_stdlib import _parse_range

    min_col, min_row, max_col, max_row = _parse_range(cell_range)
    book = xlrd.open_workbook(str(workbook_path), formatting_info=False)
    sheet_name = "Steps"
    try:
        sheet = book.sheet_by_name(sheet_name)
    except xlrd.biffh.XLRDError:
        match = next((n for n in book.sheet_names() if n.upper() == "STEPS"), None)
        if not match:
            raise RuntimeError(f"Sheet Steps not found in {workbook_path.name}") from None
        sheet = book.sheet_by_name(match)

    def _cell_value(row: int, col: int) -> Any:
        ctype = sheet.cell_type(row, col)
        val = sheet.cell_value(row, col)
        if ctype == xlrd.XL_CELL_DATE:
            from xlrd.xldate import xldate_as_datetime

            return xldate_as_datetime(val, book.datemode)
        if ctype == xlrd.XL_CELL_EMPTY:
            return None
        if ctype == xlrd.XL_CELL_TEXT:
            s = str(val).strip()
            return s or None
        return val

    return [
        [_cell_value(r, c) for c in range(min_col - 1, max_col)]
        for r in range(min_row - 1, max_row)
    ]


def _read_steps_grid_xlsx(workbook_path: Path, cell_range: str = "W2:AB10") -> List[List[Any]]:
    try:
        import openpyxl  # noqa: F401

        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            sheet_name = "Steps"
            if sheet_name not in wb.sheetnames:
                match = next((n for n in wb.sheetnames if n.upper() == "STEPS"), None)
                if not match:
                    raise RuntimeError(f"Sheet Steps not found in {workbook_path.name}")
                sheet_name = match
            ws = wb[sheet_name]
            min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(cell_range)
            return [
                list(row)
                for row in ws.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                )
            ]
        finally:
            wb.close()
    except ImportError:
        from sggg.xlsx_stdlib import read_sheet_range

        return read_sheet_range(workbook_path, "Steps", cell_range)


def _read_steps_grid(workbook_path: Path, cell_range: str = "W2:AK10") -> List[List[Any]]:
    suffix = workbook_path.suffix.lower()
    if suffix == ".xls":
        return _read_steps_grid_xls(workbook_path, cell_range)
    if suffix == ".xlsx":
        return _read_steps_grid_xlsx(workbook_path, cell_range)
    raise RuntimeError(
        f"Unsupported workbook format {suffix!r} for {workbook_path.name}; expected .xls or .xlsx"
    )


def find_compliance_workbook(
    valuation_date: date,
    root: Optional[str] = None,
) -> Tuple[Optional[Path], Optional[str]]:
    """
    Find the compliance check file for valuation_date: latest mtime among files
    whose filename date equals valuation_date. No cross-date fallback.

    Always re-scans the directory on every call. The data bridge is a
    long-running Flask process and compliance reviewers commonly save a
    midday revision and then an EOD revision later (e.g. after 5:45pm); a
    persistent module-level cache would pin the first hit and never pick
    up the later save.
    """
    base = Path(root or os.environ.get("COMPLIANCE_CHECK_ROOT", DEFAULT_ROOT))
    if not base.exists():
        return None, f"Compliance check root not found: {base}"

    # Date in filename — glob only that day (faster than rglob + parse every file).
    date_token = valuation_date.strftime("%Y.%m.%d")
    pattern = f"EHP Alt Funds - compliance check - {date_token}*"
    matches: List[Tuple[float, Path]] = []
    seen: set[str] = set()
    for path in base.glob(f"**/{pattern}"):
        if path.name.startswith("~$"):
            continue
        if path.suffix.lower() not in (".xls", ".xlsx"):
            continue
        key = str(path.resolve())
        if key in seen:
            continue
        seen.add(key)
        try:
            matches.append((path.stat().st_mtime, path))
        except OSError:
            continue

    if not matches:
        return (
            None,
            f"No compliance check workbook found for {valuation_date.isoformat()} under {base}",
        )

    matches.sort(key=lambda t: t[0], reverse=True)
    chosen = matches[0][1]
    note = None
    if len(matches) > 1:
        # Surface the chosen file's mtime so reviewers can see *which* save
        # was picked when both a midday and an EOD revision exist.
        chosen_mtime_iso = datetime.fromtimestamp(matches[0][0]).isoformat(
            timespec="seconds"
        )
        note = (
            f"{len(matches)} compliance workbooks for {valuation_date.isoformat()};"
            f" using latest save ({chosen.name} @ {chosen_mtime_iso})"
        )
    return chosen, note


def read_steps_estimates(workbook_path: Path) -> Dict[str, Dict[str, Any]]:
    """
    Parse Steps tab: W = fund, X = Current AUM, Z = Prior EOD AUM, AB = ROR, AF = Net Subs (reds).
    """
    grid = _read_steps_grid(workbook_path)
    if not grid:
        return {}

    # W2:AB10 -> row 0 is excel row 2; data funds start row 1 (excel 3)
    sheet_as_of: Optional[str] = None
    if grid and grid[0]:
        raw_date = grid[0][0]
        if raw_date is not None:
            if isinstance(raw_date, datetime):
                sheet_as_of = raw_date.date().isoformat()
            elif hasattr(raw_date, "isoformat"):
                sheet_as_of = raw_date.isoformat()[:10]
            else:
                sheet_as_of = str(raw_date).strip()[:10] or None

    out: Dict[str, Dict[str, Any]] = {}
    for row in grid[1:]:
        cells = list(row) + [None] * (10 - len(row))
        fund_raw = cells[0]
        if fund_raw is None or not str(fund_raw).strip():
            continue
        fund_label = str(fund_raw).strip()
        norm = _norm_name(fund_label)
        if norm.startswith("TOTAL"):
            continue

        fund_id = FUND_NAME_TO_ID.get(norm)
        current_aum = _parse_money(cells[1])
        prior_eod_aum = _parse_money(cells[3])
        ror_raw = cells[5]  # AB
        net_subs_reds = _parse_money(cells[9])  # AF Net Subs (reds); inflows positive, redemptions negative
        estimate_bps = _return_value_to_bps(ror_raw, prior_eod_aum)
        ror_display = str(ror_raw).strip() if ror_raw is not None else None

        entry: Dict[str, Any] = {
            "fund_label": fund_label,
            "estimate_bps": estimate_bps,
            "ror_display": ror_display,
            "current_aum": current_aum,
            "prior_eod_aum": prior_eod_aum,
            "net_subs_reds": net_subs_reds,
            "aum_currency": FUND_NATIVE_CURRENCY.get(fund_id or "", "CAD"),
            "fund_id": fund_id,
        }
        if fund_id:
            out[fund_id] = entry
        else:
            out[norm] = entry

    meta = {"sheet_as_of": sheet_as_of}
    return {"by_fund_id": {k: v for k, v in out.items() if len(k) == 36 and "-" in k}, "meta": meta}


def compliance_aum_change_ex_flows(
    prior_eod_aum: Optional[float],
    current_aum: Optional[float],
    net_subs_reds: Optional[float],
) -> Optional[float]:
    """
    Dollar change aligned with Steps column AB: (X - Z) minus AF net subs/reds.
    Redemptions are negative in AF; subtracting AF removes flow from the raw AUM delta.
    """
    if prior_eod_aum is None or current_aum is None:
        return None
    delta = float(current_aum) - float(prior_eod_aum)
    if net_subs_reds is None:
        return delta
    return delta - float(net_subs_reds)


def estimates_by_fund_id(
    valuation_date: date,
    root: Optional[str] = None,
) -> Dict[str, Any]:
    """Resolve workbook and return estimates keyed by fund GUID."""
    path, note = find_compliance_workbook(valuation_date, root=root)
    if not path:
        return {
            "available": False,
            "error": note or "Workbook not found",
            "estimates_by_fund_id": {},
            "file_variant": None,
            "saved_at": None,
        }
    try:
        mtime = path.stat().st_mtime
        saved_at = datetime.fromtimestamp(mtime).isoformat(timespec="seconds")
    except OSError:
        saved_at = None

    try:
        parsed = read_steps_estimates(path)
        by_fund = parsed.get("by_fund_id") or {}
        meta = parsed.get("meta") or {}
    except Exception as e:
        return {
            "available": False,
            "error": str(e),
            "workbook_path": str(path),
            "estimates_by_fund_id": {},
            "file_variant": _file_variant(path),
            "saved_at": saved_at,
        }

    estimates: Dict[str, Dict[str, Any]] = {}
    for fid, row in by_fund.items():
        estimates[fid] = {
            **row,
            "spreadsheet_label": row.get("fund_label"),
        }

    sheet_as_of = meta.get("sheet_as_of")
    date_warning = None
    if sheet_as_of and sheet_as_of != valuation_date.isoformat():
        date_warning = f"Steps tab date {sheet_as_of} differs from valuation date {valuation_date.isoformat()}"

    return {
        "available": True,
        "workbook_path": str(path),
        "note": note,
        "date_warning": date_warning,
        "sheet_as_of": sheet_as_of,
        "file_variant": _file_variant(path),
        "saved_at": saved_at,
        "estimates_by_fund_id": estimates,
    }
